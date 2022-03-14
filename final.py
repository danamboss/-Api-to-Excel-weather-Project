import os
from openpyxl import load_workbook
import openpyxl
import requests
import json 
from openpyxl import Workbook
from openpyxl.chart import (
    Reference,
    Series,
    BarChart3D,
)

def read_excel(name):
  wb = load_workbook(name)
  
  print(type(wb)) 
  print(wb.sheetnames) 
  
  sheet1 = wb["Sheet1"] 
  sheet2 = wb["Sheet2"]  
  sheet3 = wb["Sheet3"]
  
  city1 = sheet1["A1"].value
  city2 = sheet2["A1"].value
  
  cities = [city1, city2]
  filename = 'data.json'
  cities_name = []
  
  for city in cities:
    my_secret = os.environ['keyw']
    url = f'https://api.openweathermap.org/data/2.5/weather?q={city},CA&appid={my_secret}'

    response = requests.get(url)
    
    response_dict = response.json()
    cities_name.append(response_dict)
  with open(filename, 'w') as file_object:
    json.dump(cities_name, file_object, indent = 4)
  
  with open(filename) as file_object:
    myweather = json.load(file_object)
  
  temperatures, humidity = [], []
  for weather in myweather:
    temp = weather['main']['temp']
    hum = weather['main']['humidity']
    temperatures.append(temp)
    humidity.append(hum)
  
  temp_1 = round(temperatures[0]-273.15, 2)
  temp_2 = round(temperatures[1]-273.15, 2)
  print(temp_1)
  print(temp_2)
  celcius_temp = []
  celcius_temp.append(temp_1)
  celcius_temp.append(temp_2)
  print(celcius_temp)
  
  rows = [
       (None, None),
      ("Temperature",celcius_temp[0]),
      ("Humidity", humidity[0]),
      
  ]
  
  for row in rows:
      sheet1.append(row)
  
  data = Reference(sheet1, min_col=2, min_row=1, max_col=3, max_row=4)
  titles = Reference(sheet1, min_col=1, min_row=2, max_row=4)
  chart = BarChart3D()
  chart.title = "Bar Chart for Temperature and Humidity"
  chart.add_data(data=data, titles_from_data=True)
  chart.set_categories(titles)
  chart.x_axis.title = city1
  chart.y_axis.title = "Temperature and Humidity"
  sheet1.add_chart(chart, "E5")
  
  rows = [
      (None, None),
      ("Temperature",celcius_temp[1]),
      ("Humidity", humidity[1]),
        
  ]
  
  for row in rows:
      sheet2.append(row)
  
  data = Reference(sheet2, min_col=2, min_row=1, max_col=3, max_row=4)
  titles = Reference(sheet2, min_col=1, min_row=2, max_row=4)
  chart = BarChart3D()
  chart.title = "Bar Chart for Temperature and Humidity"
  chart.add_data(data=data, titles_from_data=True)
  chart.set_categories(titles)
  chart.x_axis.title = city2
  chart.y_axis.title = "Temperature and Humidity"
  sheet2.add_chart(chart, "E5")

  rows = [
      ('City', 'Temperature' , 'Humidity'),
      (city1,celcius_temp[0] , humidity[0]),
      (city2, celcius_temp[1], humidity[1]),
      
  ]
  
  for row in rows:
      sheet3.append(row)
  
  data = Reference(sheet3, min_col=2, min_row=1, max_col=3, max_row=4)
  titles = Reference(sheet3, min_col=1, min_row=2, max_row=4)
  chart = BarChart3D()
  chart.title = "Comparison between Temperature and Humidity"
  chart.add_data(data=data, titles_from_data=True)
  chart.set_categories(titles)
  chart.x_axis.title = "Canada Cities"
  chart.y_axis.title = "Temperature and Humidity"
  sheet3.add_chart(chart, "E5")
  
  

  wb.save('static/processed/weather.xlsx')